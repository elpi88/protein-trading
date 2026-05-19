"""
Layer di accesso al file Excel.
Tutta la lettura/scrittura passa di qui.

Il file Excel rimane la fonte di verita'. L'app fa un backup automatico
prima di ogni scrittura, dentro ../backups/.
"""
from __future__ import annotations

import re
import shutil
from datetime import datetime
from pathlib import Path
from typing import Optional

import openpyxl
import pandas as pd
import streamlit as st

# ----------------------------------------------------------------------
# Percorsi
# ----------------------------------------------------------------------
APP_DIR = Path(__file__).resolve().parent.parent
PROJECT_DIR = APP_DIR.parent
EXCEL_FILE = PROJECT_DIR / "Protein_Trading_ERP_FULL.xlsm"
BACKUP_DIR = PROJECT_DIR / "backups"
ATTACH_DIR = PROJECT_DIR / "attachments"
BACKUP_DIR.mkdir(exist_ok=True)
ATTACH_DIR.mkdir(exist_ok=True)

# ----------------------------------------------------------------------
# Schema dei fogli: lista colonne in ordine (per scrittura)
# ----------------------------------------------------------------------
SCHEMAS = {
    "SUPPLIERS_CLEAN": [
        "Supplier ID", "Company Name", "Contact Person", "Protein Category",
        "Country", "Email", "Phone", "Website", "Address", "Products",
        "Tax/VAT", "Registration", "Notes",
    ],
    "CLIENTS": [
        "Client ID", "Company Name", "CONTACT PERSON", "PROTEIN CATEGORY",
        "ITEMS", "COUNTRY", "Email", "Phone", "Monthly Capacity", "Notes",
    ],
    "OFFERS": [
        "Offer ID", "Supplier", "Product", "Subproduct", "Specifics",
        "Packaging", "Price", "Currency", "Unit", "Price USD/kg", "Incoterm",
        "Country Destination", "Load Ready Date", "Offer Date", "Source",
        "Notes", "Match Key",
    ],
    "BIDS": [
        "Bid ID", "Client", "Product", "Subproduct", "Specifics", "Packaging",
        "Target Price", "Currency", "Unit", "Target USD/kg", "Volume (kg)",
        "Incoterm", "Origin Country", "Need By Date", "Bid Date", "Status",
        "Notes",
    ],
    "SHIPMENTS": [
        "Shipment ID", "Invoice ID", "Bid ID", "Client", "Product",
        "Quantity", "Unit", "Origin Port", "Destination Port",
        "Carrier/Vessel", "Container #", "Container Type", "Incoterm",
        "ETD", "ETA", "Days in Transit", "Status", "Tracking #", "Notes",
    ],
    "INVOICES": [
        "Invoice ID", "Date", "Bid ID", "Client", "Product",
        "Quantity", "Unit", "Unit Price", "Currency",
        "Subtotal", "Subtotal USD", "VAT %", "VAT Amount",
        "Total Invoice", "Total USD", "Payment Status",
        "Due Date", "Paid Date", "Notes",
    ],
}

ID_PREFIX = {
    "SUPPLIERS_CLEAN": "SUP",
    "CLIENTS": "CLI",
    "OFFERS": "OFF",
    "BIDS": "BID",
    "SHIPMENTS": "SHP",
    "INVOICES": "INV",
}

# Colonne calcolate (formule Excel) - non vanno mai sovrascritte
FORMULA_COLS = {
    "OFFERS": [10, 17],   # Price USD/kg, Match Key  (1-based)
    "BIDS": [10],         # Target USD/kg
}


# ----------------------------------------------------------------------
# Helper: backup
# ----------------------------------------------------------------------
def make_backup() -> Path:
    """Backup del file Excel prima di una scrittura. Tiene gli ultimi 20."""
    ts = datetime.now().strftime("%Y%m%d_%H%M%S")
    dest = BACKUP_DIR / f"backup_{ts}_Protein_Trading_ERP_FULL.xlsm"
    shutil.copy(EXCEL_FILE, dest)
    # ruota: tieni solo gli ultimi 20
    backups = sorted(BACKUP_DIR.glob("backup_*.xlsm"))
    for old in backups[:-20]:
        try:
            old.unlink()
        except Exception:
            pass
    return dest


# ----------------------------------------------------------------------
# Lettura
# ----------------------------------------------------------------------
@st.cache_data(show_spinner=False, ttl=2)
def read_sheet(sheet_name: str) -> pd.DataFrame:
    """Legge un foglio Excel in DataFrame, valori calcolati (data_only)."""
    df = pd.read_excel(EXCEL_FILE, sheet_name=sheet_name, engine="openpyxl")
    # rimuovi righe completamente vuote
    df = df.dropna(how="all")
    # normalizza colonna ID come stringa
    if len(df.columns) > 0:
        df[df.columns[0]] = df[df.columns[0]].astype(str).replace("nan", "")
        df = df[df[df.columns[0]] != ""]
    return df.reset_index(drop=True)


def clear_cache():
    """Forza il refresh dei dati al prossimo read_sheet."""
    read_sheet.clear()


# ----------------------------------------------------------------------
# ID generation
# ----------------------------------------------------------------------
def next_id(sheet_name: str) -> str:
    """Prossimo ID = MAX numero esistente + 1 (NON la posizione di riga)."""
    prefix = ID_PREFIX[sheet_name]
    df = read_sheet(sheet_name)
    if df.empty:
        return f"{prefix}-00001"
    max_num = 0
    for v in df[df.columns[0]].astype(str):
        m = re.match(rf"{prefix}-?(\d+)", v, re.IGNORECASE)
        if m:
            max_num = max(max_num, int(m.group(1)))
    return f"{prefix}-{max_num + 1:05d}"


# ----------------------------------------------------------------------
# Scrittura (CREATE / UPDATE / DELETE)
# ----------------------------------------------------------------------
def add_row(sheet_name: str, values: dict) -> str:
    """
    Aggiunge una riga al foglio. `values` è un dict colonna->valore (senza ID).
    Ritorna il nuovo ID generato.
    Le colonne formula vengono scritte come formule Excel.
    """
    new_id = next_id(sheet_name)
    make_backup()

    wb = openpyxl.load_workbook(EXCEL_FILE, keep_vba=True)
    ws = wb[sheet_name]
    columns = SCHEMAS[sheet_name]
    formula_cols = FORMULA_COLS.get(sheet_name, [])

    # trova prima riga vuota
    next_row = ws.max_row + 1
    for r in range(2, ws.max_row + 1):
        if ws.cell(r, 1).value is None or str(ws.cell(r, 1).value).strip() == "":
            next_row = r
            break

    ws.cell(next_row, 1).value = new_id
    for idx, col in enumerate(columns[1:], start=2):
        if idx in formula_cols:
            if sheet_name == "OFFERS" and idx == 10:
                ws.cell(next_row, 10).value = (
                    f"=IFERROR(G{next_row}*VLOOKUP(H{next_row},FX_TABLE,3,FALSE())"
                    f"/VLOOKUP(I{next_row},WT_TABLE,3,FALSE()),\"\")"
                )
            elif sheet_name == "OFFERS" and idx == 17:
                ws.cell(next_row, 17).value = (
                    f'=C{next_row}&"|"&TEXT(J{next_row},"0.0000")'
                )
            elif sheet_name == "BIDS" and idx == 10:
                ws.cell(next_row, 10).value = (
                    f"=IFERROR(G{next_row}*VLOOKUP(H{next_row},FX_TABLE,3,FALSE())"
                    f"/VLOOKUP(I{next_row},WT_TABLE,3,FALSE()),\"\")"
                )
            continue
        ws.cell(next_row, idx).value = values.get(col)

    wb.save(EXCEL_FILE)
    clear_cache()
    try: log_action("ADD", sheet_name, new_id, str(values.get("Company Name") or values.get("Supplier") or values.get("Client") or "")[:200])
    except Exception: pass
    return new_id


def update_row(sheet_name: str, row_id: str, values: dict) -> bool:
    """Aggiorna la riga con ID = row_id."""
    make_backup()

    wb = openpyxl.load_workbook(EXCEL_FILE, keep_vba=True)
    ws = wb[sheet_name]
    columns = SCHEMAS[sheet_name]
    formula_cols = FORMULA_COLS.get(sheet_name, [])

    target = None
    for r in range(2, ws.max_row + 1):
        if str(ws.cell(r, 1).value).strip().upper() == row_id.strip().upper():
            target = r
            break
    if target is None:
        return False

    for idx, col in enumerate(columns[1:], start=2):
        if idx in formula_cols:
            continue  # non toccare le formule
        ws.cell(target, idx).value = values.get(col)

    wb.save(EXCEL_FILE)
    clear_cache()
    try: log_action("UPDATE", sheet_name, row_id, str(values.get("Company Name") or values.get("Supplier") or values.get("Client") or "")[:200])
    except Exception: pass
    return True


def delete_row(sheet_name: str, row_id: str) -> bool:
    """Cancella la riga con ID = row_id (delete fisico, riga rimossa)."""
    make_backup()
    wb = openpyxl.load_workbook(EXCEL_FILE, keep_vba=True)
    ws = wb[sheet_name]
    target = None
    for r in range(2, ws.max_row + 1):
        if str(ws.cell(r, 1).value).strip().upper() == row_id.strip().upper():
            target = r
            break
    if target is None:
        return False
    ws.delete_rows(target)
    wb.save(EXCEL_FILE)
    clear_cache()
    try: log_action("DELETE", sheet_name, row_id, "")
    except Exception: pass
    return True


# ----------------------------------------------------------------------
# Liste di riferimento (drop-down) dal foglio REFERENCE_LISTS
# ----------------------------------------------------------------------
@st.cache_data(show_spinner=False, ttl=60)
def get_protein_categories() -> list[str]:
    try:
        df = pd.read_excel(EXCEL_FILE, sheet_name="REFERENCE_LISTS",
                           engine="openpyxl", header=None)
        # PROTEIN_LIST è in colonna A righe 4..N (riga 3 = header)
        vals = df.iloc[3:, 0].dropna().astype(str).tolist()
        return [v for v in vals if v.strip()]
    except Exception:
        return ["FISH", "PORK", "BEEF", "POULTRY", "LAMB", "TRADER",
                "POTATOES", "OTHER", "UNCLASSIFIED"]


@st.cache_data(show_spinner=False, ttl=60)
def get_countries() -> list[str]:
    try:
        df = pd.read_excel(EXCEL_FILE, sheet_name="REFERENCE_LISTS",
                           engine="openpyxl", header=None)
        vals = df.iloc[3:, 2].dropna().astype(str).tolist()
        return [v for v in vals if v.strip()]
    except Exception:
        return ["Switzerland", "Italy", "Germany", "Other"]


@st.cache_data(show_spinner=False, ttl=60)
def get_currencies() -> list[str]:
    try:
        df = pd.read_excel(EXCEL_FILE, sheet_name="CONVERSIONS",
                           engine="openpyxl", header=None)
        # FX_CODES = A6:A25 (riga 6 in Excel = indice 5)
        vals = df.iloc[5:25, 0].dropna().astype(str).tolist()
        return [v for v in vals if v.strip()]
    except Exception:
        return ["USD", "EUR", "CHF", "GBP", "CNY"]


@st.cache_data(show_spinner=False, ttl=60)
def get_units() -> list[str]:
    try:
        df = pd.read_excel(EXCEL_FILE, sheet_name="CONVERSIONS",
                           engine="openpyxl", header=None)
        vals = df.iloc[29:36, 0].dropna().astype(str).tolist()
        return [v for v in vals if v.strip()]
    except Exception:
        return ["KG", "LB", "MT"]


# ----------------------------------------------------------------------
# Statistiche per Dashboard
# ----------------------------------------------------------------------
def get_kpis() -> dict:
    """Numeri sintetici per la home."""
    s = read_sheet("SUPPLIERS_CLEAN")
    c = read_sheet("CLIENTS")
    o = read_sheet("OFFERS")
    b = read_sheet("BIDS")

    out = {
        "n_suppliers": len(s),
        "n_clients": len(c),
        "n_offers": len(o),
        "n_bids": len(b),
    }
    # bid aperti
    if "Status" in b.columns:
        out["n_bids_open"] = int((b["Status"].astype(str).str.upper() == "OPEN").sum())
    else:
        out["n_bids_open"] = 0

    # valore stimato pipeline = somma (Target USD/kg * Volume kg) dei bid OPEN
    try:
        b_open = b[b["Status"].astype(str).str.upper() == "OPEN"].copy()
        b_open["Target USD/kg"] = pd.to_numeric(b_open["Target USD/kg"], errors="coerce")
        # se la formula non e' calcolata, ricalcola noi
        miss = b_open["Target USD/kg"].isna()
        if miss.any():
            b_open.loc[miss, "Target USD/kg"] = b_open[miss].apply(
                lambda r: compute_usd_kg(r.get("Target Price"), r.get("Currency"), r.get("Unit")),
                axis=1
            )
        b_open["Volume (kg)"] = pd.to_numeric(b_open["Volume (kg)"], errors="coerce")
        out["pipeline_usd"] = float((b_open["Target USD/kg"] * b_open["Volume (kg)"]).sum(skipna=True))
    except Exception:
        out["pipeline_usd"] = 0.0
    return out


# ======================================================================
# FX / WT lookup - calcolo USD/kg in Python quando la formula non e' valutata
# ======================================================================
@st.cache_data(show_spinner=False, ttl=300)
def get_fx_table() -> dict[str, float]:
    """Ritorna {currency_code: rate_to_usd_per_unit}. Legge FX_TABLE da CONVERSIONS A6:C25."""
    try:
        df = pd.read_excel(EXCEL_FILE, sheet_name="CONVERSIONS",
                           engine="openpyxl", header=None)
        # A6:C25 = righe 6..25 in Excel = indici 5..24
        out = {}
        for i in range(5, 25):
            code = df.iat[i, 0]
            rate = df.iat[i, 2]
            if pd.notna(code) and pd.notna(rate):
                try:
                    out[str(code).strip().upper()] = float(rate)
                except Exception:
                    pass
        return out
    except Exception:
        return {"USD": 1.0}


@st.cache_data(show_spinner=False, ttl=300)
def get_wt_table() -> dict[str, float]:
    """Ritorna {unit_code: kg_factor}. Legge WT_TABLE da CONVERSIONS A30:C36."""
    try:
        df = pd.read_excel(EXCEL_FILE, sheet_name="CONVERSIONS",
                           engine="openpyxl", header=None)
        out = {}
        for i in range(29, 36):
            code = df.iat[i, 0]
            factor = df.iat[i, 2]
            if pd.notna(code) and pd.notna(factor):
                try:
                    out[str(code).strip().upper()] = float(factor)
                except Exception:
                    pass
        return out
    except Exception:
        return {"KG": 1.0, "LB": 0.453592, "MT": 1000.0}


def compute_usd_kg(price, currency, unit) -> Optional[float]:
    """Replica della formula Excel: =price * FX[curr] / WT[unit]."""
    try:
        p = float(price)
        if p <= 0:
            return None
        fx = get_fx_table().get(str(currency).strip().upper())
        wt = get_wt_table().get(str(unit).strip().upper())
        if not fx or not wt:
            return None
        return p * fx / wt
    except Exception:
        return None


# ======================================================================
# Matching offerte ↔ bid
# ======================================================================
def normalize_product(s) -> str:
    """Normalizza nomi prodotto per matching: upper, trim, collassa spazi."""
    if not s or pd.isna(s):
        return ""
    return " ".join(str(s).upper().split())


def _ensure_usd_kg(df: pd.DataFrame, price_col: str, usd_kg_col: str) -> pd.DataFrame:
    """Garantisce che usd_kg_col abbia valori; se mancano, ricalcola."""
    df = df.copy()
    df[usd_kg_col] = pd.to_numeric(df.get(usd_kg_col), errors="coerce")
    miss = df[usd_kg_col].isna()
    if miss.any() and price_col in df.columns:
        df.loc[miss, usd_kg_col] = df.loc[miss].apply(
            lambda r: compute_usd_kg(r.get(price_col), r.get("Currency"), r.get("Unit")),
            axis=1
        )
    return df


def get_matches(open_bids_only: bool = True) -> pd.DataFrame:
    """
    Ritorna tutti i match offerta↔bid basati su Product (normalizzato uguale).
    Colonne: Bid ID, Client, Offer ID, Supplier, Product, Subproduct,
             Target USD/kg, Offer USD/kg, Margin USD/kg, Volume (kg),
             Margin USD (totale), Status, Need By Date, Offer Date
    """
    offers = read_sheet("OFFERS")
    bids = read_sheet("BIDS")
    if offers.empty or bids.empty:
        return pd.DataFrame()

    offers = _ensure_usd_kg(offers, "Price", "Price USD/kg")
    bids = _ensure_usd_kg(bids, "Target Price", "Target USD/kg")

    if open_bids_only and "Status" in bids.columns:
        bids = bids[bids["Status"].astype(str).str.upper() == "OPEN"].copy()

    offers["_pkey"] = offers["Product"].apply(normalize_product)
    bids["_pkey"] = bids["Product"].apply(normalize_product)

    merged = bids.merge(offers, on="_pkey", how="inner",
                         suffixes=(" (bid)", " (offer)"))
    if merged.empty:
        return pd.DataFrame()

    merged["Margin USD/kg"] = (
        pd.to_numeric(merged["Target USD/kg"], errors="coerce")
        - pd.to_numeric(merged["Price USD/kg"], errors="coerce")
    )
    merged["Volume (kg)"] = pd.to_numeric(merged["Volume (kg)"], errors="coerce").fillna(0)
    merged["Margin USD"] = merged["Margin USD/kg"] * merged["Volume (kg)"]

    cols = [
        "Bid ID", "Client", "Offer ID", "Supplier",
        "Product (bid)", "Subproduct (bid)", "Subproduct (offer)",
        "Target USD/kg", "Price USD/kg", "Margin USD/kg",
        "Volume (kg)", "Margin USD",
        "Status", "Need By Date", "Offer Date",
    ]
    cols = [c for c in cols if c in merged.columns]
    return merged[cols].sort_values("Margin USD/kg", ascending=False, na_position="last").reset_index(drop=True)


# ======================================================================
# Allegati
# ======================================================================
def attachments_dir(sheet: str, row_id: str) -> Path:
    """Cartella allegati per un singolo record."""
    d = ATTACH_DIR / sheet / row_id
    d.mkdir(parents=True, exist_ok=True)
    return d


def list_attachments(sheet: str, row_id: str) -> list[Path]:
    """File allegati per un record."""
    d = ATTACH_DIR / sheet / row_id
    if not d.exists():
        return []
    return sorted([p for p in d.iterdir() if p.is_file()])


def save_attachment(sheet: str, row_id: str, filename: str, content: bytes) -> Path:
    """Salva un file negli allegati del record. Sanifica il nome file."""
    safe = "".join(c for c in filename if c.isalnum() or c in "._- ()[]").strip()
    if not safe:
        safe = "file"
    d = attachments_dir(sheet, row_id)
    dest = d / safe
    # se gia' esiste, append timestamp
    if dest.exists():
        from datetime import datetime
        ts = datetime.now().strftime("%Y%m%d_%H%M%S")
        stem = dest.stem
        suf = dest.suffix
        dest = d / f"{stem}_{ts}{suf}"
    dest.write_bytes(content)
    return dest


def delete_attachment(sheet: str, row_id: str, filename: str) -> bool:
    p = ATTACH_DIR / sheet / row_id / filename
    if p.exists() and p.is_file():
        p.unlink()
        return True
    return False


# ======================================================================
# AUDIT LOG (storico modifiche)
# File CSV separato: NON intacca le macro Excel.
# ======================================================================
import csv

AUDIT_LOG = PROJECT_DIR / "audit_log.csv"
AUDIT_COLUMNS = ["timestamp", "user", "action", "sheet", "row_id", "details"]


def _ensure_audit_file():
    if not AUDIT_LOG.exists():
        with open(AUDIT_LOG, "w", newline="", encoding="utf-8") as f:
            csv.writer(f).writerow(AUDIT_COLUMNS)


def log_action(action: str, sheet: str, row_id: str = "", details: str = ""):
    """Loga un'azione nell'audit log. user è preso da session_state se presente."""
    _ensure_audit_file()
    user = "system"
    try:
        user = st.session_state.get("user", "system")
    except Exception:
        pass
    with open(AUDIT_LOG, "a", newline="", encoding="utf-8") as f:
        csv.writer(f).writerow([
            datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
            user, action, sheet, row_id, details[:500]
        ])


def read_audit_log(limit: int = 500) -> pd.DataFrame:
    _ensure_audit_file()
    df = pd.read_csv(AUDIT_LOG)
    if len(df) > limit:
        df = df.tail(limit)
    return df.iloc[::-1].reset_index(drop=True)  # ultime prima


# ======================================================================
# MERGE record duplicati
# ======================================================================
def merge_records(sheet: str, keep_id: str, drop_id: str, chosen: dict[str, str]) -> bool:
    """
    Fonde drop_id dentro keep_id:
    - chosen: dict colonna->"keep"|"drop"|valore_custom: scegli quale valore mantenere per ogni colonna
    - Cancella drop_id
    - Aggiorna i riferimenti in OFFERS.Supplier o BIDS.Client se sheet è SUPPLIERS_CLEAN/CLIENTS
    """
    make_backup()
    wb = openpyxl.load_workbook(EXCEL_FILE, keep_vba=True)
    ws = wb[sheet]
    columns = SCHEMAS[sheet]
    formula_cols = FORMULA_COLS.get(sheet, [])

    # Trova le righe
    keep_row = drop_row = None
    for r in range(2, ws.max_row + 1):
        v = str(ws.cell(r, 1).value).strip().upper()
        if v == keep_id.strip().upper(): keep_row = r
        elif v == drop_id.strip().upper(): drop_row = r
    if keep_row is None or drop_row is None:
        return False

    # Determina i nomi azienda prima e dopo per il riferimento in OFFERS/BIDS
    old_name = str(ws.cell(drop_row, 2).value or "")
    keep_name_before = str(ws.cell(keep_row, 2).value or "")

    # Applica le scelte sulle colonne (col 2..N)
    for idx, col in enumerate(columns[1:], start=2):
        if idx in formula_cols:
            continue
        choice = chosen.get(col, "keep")
        if choice == "keep":
            pass  # niente da fare
        elif choice == "drop":
            ws.cell(keep_row, idx).value = ws.cell(drop_row, idx).value
        else:
            ws.cell(keep_row, idx).value = choice  # valore custom

    keep_name_after = str(ws.cell(keep_row, 2).value or "")

    # Cancella la riga drop
    ws.delete_rows(drop_row)

    # Aggiorna riferimenti nominali in OFFERS / BIDS
    refs_updated = 0
    if sheet == "SUPPLIERS_CLEAN":
        ws_off = wb["OFFERS"]
        for r in range(2, ws_off.max_row + 1):
            cell = ws_off.cell(r, 2)  # col Supplier
            v = str(cell.value or "")
            if v == old_name and old_name != keep_name_after:
                cell.value = keep_name_after
                refs_updated += 1
    elif sheet == "CLIENTS":
        ws_bid = wb["BIDS"]
        for r in range(2, ws_bid.max_row + 1):
            cell = ws_bid.cell(r, 2)  # col Client
            v = str(cell.value or "")
            if v == old_name and old_name != keep_name_after:
                cell.value = keep_name_after
                refs_updated += 1
        # SHIPMENTS e INVOICES hanno anche col Client (4)
        for sname in ("SHIPMENTS", "INVOICES"):
            if sname in wb.sheetnames:
                wss = wb[sname]
                for r in range(2, wss.max_row + 1):
                    cell = wss.cell(r, 4)
                    v = str(cell.value or "")
                    if v == old_name and old_name != keep_name_after:
                        cell.value = keep_name_after
                        refs_updated += 1

    wb.save(EXCEL_FILE)
    clear_cache()
    log_action("MERGE", sheet, keep_id,
                 f"merged {drop_id} into {keep_id}; refs updated: {refs_updated}")
    return True


# ======================================================================
# Trova duplicati potenziali
# ======================================================================
def find_potential_duplicates(sheet: str) -> pd.DataFrame:
    """
    Trova coppie di record con nome MOLTO simile (case insensitive,
    ignorando spazi/punteggiatura). Ritorna DataFrame con coppie candidate.
    """
    df = read_sheet(sheet)
    if df.empty or "Company Name" not in df.columns:
        return pd.DataFrame()
    id_col = df.columns[0]

    def norm(s):
        if not s or pd.isna(s): return ""
        return "".join(c.lower() for c in str(s) if c.isalnum())

    df = df.copy()
    df["_norm"] = df["Company Name"].apply(norm)
    df = df[df["_norm"] != ""]
    grp = df.groupby("_norm").filter(lambda g: len(g) >= 2)
    if grp.empty:
        return pd.DataFrame()

    pairs = []
    for _, g in grp.groupby("_norm"):
        rows = g.to_dict("records")
        for i in range(len(rows)):
            for j in range(i + 1, len(rows)):
                pairs.append({
                    "Keep candidate": rows[i][id_col],
                    "Duplicate candidate": rows[j][id_col],
                    "Name (keep)": rows[i]["Company Name"],
                    "Name (dup)": rows[j]["Company Name"],
                })
    return pd.DataFrame(pairs)


# ======================================================================
# Fuzzy matching duplicati + merge con CONCATENAZIONE
# ======================================================================
from difflib import SequenceMatcher

# Suffissi legali da rimuovere per il confronto fuzzy
_LEGAL_SUFFIXES = {
    "srl", "s.r.l", "s.r.l.", "spa", "s.p.a", "s.p.a.", "sa", "s.a", "s.a.",
    "ag", "gmbh", "g.m.b.h", "ltd", "ltd.", "limited", "inc", "inc.",
    "llc", "l.l.c", "l.l.c.", "co", "co.", "corp", "corp.", "corporation",
    "company", "trading", "trade", "international", "intl", "group",
    "holding", "holdings", "bv", "b.v", "b.v.", "nv", "n.v", "n.v.",
    "kg", "ohg", "se", "plc", "p.l.c",
}


def _norm_company(s) -> str:
    """Normalizza un nome azienda per il confronto fuzzy:
       lowercase, rimuove punti (per riconoscere S.R.L. = SRL),
       poi rimuove punteggiatura residua, suffissi legali e doppi spazi."""
    if not s or pd.isna(s):
        return ""
    txt = str(s).lower()
    # PRIMA rimuovi i punti (così S.R.L. → SRL, A.G. → AG, G.M.B.H. → GMBH)
    txt = txt.replace(".", "")
    # Poi sostituisci la restante punteggiatura con spazio
    txt = re.sub(r"[^a-z0-9\s]", " ", txt)
    # Splitta, filtra suffissi legali, ricomponi
    tokens = [t for t in txt.split() if t and t not in _LEGAL_SUFFIXES]
    return " ".join(tokens)


def _similarity(a: str, b: str) -> float:
    """Ritorna similarità 0..1 tra due stringhe (dopo normalizzazione)."""
    na, nb = _norm_company(a), _norm_company(b)
    if not na or not nb:
        return 0.0
    return SequenceMatcher(None, na, nb).ratio()


def find_fuzzy_duplicates(sheet: str, threshold: float = 0.85) -> pd.DataFrame:
    """
    Trova coppie di record con nome SIMILE sopra una soglia (0..1).
    Soglia 1.0 = nome identico (dopo normalizzazione).
    Soglia 0.85 = molto simile (default consigliato).
    Soglia 0.70 = somiglianza generica (più falsi positivi).
    Ritorna DataFrame con coppie ordinate per similarità decrescente.
    """
    df = read_sheet(sheet)
    if df.empty or "Company Name" not in df.columns:
        return pd.DataFrame()
    id_col = df.columns[0]

    # Estraggo (id, nome, nome_norm) per ridurre lavoro
    items = []
    for _, r in df.iterrows():
        nid = str(r[id_col])
        name = str(r["Company Name"]) if pd.notna(r["Company Name"]) else ""
        if not name.strip():
            continue
        items.append((nid, name, _norm_company(name)))

    pairs = []
    n = len(items)
    for i in range(n):
        id_a, name_a, norm_a = items[i]
        if not norm_a:
            continue
        for j in range(i + 1, n):
            id_b, name_b, norm_b = items[j]
            if not norm_b:
                continue
            sim = SequenceMatcher(None, norm_a, norm_b).ratio()
            if sim >= threshold:
                pairs.append({
                    "ID 1": id_a,
                    "Nome 1": name_a,
                    "ID 2": id_b,
                    "Nome 2": name_b,
                    "Similarità %": round(sim * 100, 1),
                })

    if not pairs:
        return pd.DataFrame()
    out = pd.DataFrame(pairs).sort_values("Similarità %", ascending=False).reset_index(drop=True)
    return out


def _concat_values(v1, v2, sep: str = "; ") -> str:
    """Unisce due valori in stringa, evitando duplicati e svuotando i None."""
    def is_empty(x):
        return x is None or (isinstance(x, float) and pd.isna(x)) or \
               (isinstance(x, str) and not x.strip())

    if is_empty(v1) and is_empty(v2):
        return None
    if is_empty(v1):
        return str(v2).strip()
    if is_empty(v2):
        return str(v1).strip()

    s1 = str(v1).strip()
    s2 = str(v2).strip()
    if s1.lower() == s2.lower():
        return s1

    # Splitta gli eventuali valori già concatenati e dedup case-insensitive
    parts1 = [p.strip() for p in s1.split(sep) if p.strip()]
    parts2 = [p.strip() for p in s2.split(sep) if p.strip()]
    seen_lower = set()
    out = []
    for p in parts1 + parts2:
        if p.lower() not in seen_lower:
            seen_lower.add(p.lower())
            out.append(p)
    return sep.join(out)


def merge_records_concat(sheet: str, keep_id: str, drop_id: str) -> dict:
    """
    Fonde drop_id dentro keep_id concatenando i valori dei campi con ';'.
    - Per Company Name: tiene il valore di keep_id (mai concatena il nome).
    - Per gli altri campi: concatena con `; ` solo se diversi.
    - Cancella drop_id e aggiorna riferimenti in OFFERS/BIDS/SHIPMENTS/INVOICES.
    Ritorna dict con info sul merge: {fields_merged: [...], refs_updated: N}
    """
    make_backup()
    wb = openpyxl.load_workbook(EXCEL_FILE, keep_vba=True)
    ws = wb[sheet]
    columns = SCHEMAS[sheet]
    formula_cols = FORMULA_COLS.get(sheet, [])

    # Trova le righe
    keep_row = drop_row = None
    for r in range(2, ws.max_row + 1):
        v = str(ws.cell(r, 1).value).strip().upper()
        if v == keep_id.strip().upper():
            keep_row = r
        elif v == drop_id.strip().upper():
            drop_row = r
    if keep_row is None or drop_row is None:
        return {"ok": False, "error": "ID non trovato"}

    old_name = str(ws.cell(drop_row, 2).value or "")
    keep_name = str(ws.cell(keep_row, 2).value or "")

    fields_merged = []
    # Itera sulle colonne (skip ID = col 1)
    for idx, col in enumerate(columns[1:], start=2):
        if idx in formula_cols:
            continue
        v_keep = ws.cell(keep_row, idx).value
        v_drop = ws.cell(drop_row, idx).value

        # Per il nome azienda (col 2) tieni quello del keep
        if idx == 2:
            continue

        merged = _concat_values(v_keep, v_drop)
        if merged != v_keep:
            ws.cell(keep_row, idx).value = merged
            fields_merged.append(col)

    # Cancella la riga drop
    ws.delete_rows(drop_row)

    # Aggiorna riferimenti nominali in OFFERS / BIDS / SHIPMENTS / INVOICES
    refs_updated = 0
    if sheet == "SUPPLIERS_CLEAN" and old_name and old_name != keep_name:
        if "OFFERS" in wb.sheetnames:
            ws_off = wb["OFFERS"]
            for r in range(2, ws_off.max_row + 1):
                cell = ws_off.cell(r, 2)  # col Supplier
                if str(cell.value or "") == old_name:
                    cell.value = keep_name
                    refs_updated += 1
    elif sheet == "CLIENTS" and old_name and old_name != keep_name:
        if "BIDS" in wb.sheetnames:
            ws_bid = wb["BIDS"]
            for r in range(2, ws_bid.max_row + 1):
                cell = ws_bid.cell(r, 2)  # col Client
                if str(cell.value or "") == old_name:
                    cell.value = keep_name
                    refs_updated += 1
        for sname in ("SHIPMENTS", "INVOICES"):
            if sname in wb.sheetnames:
                wss = wb[sname]
                for r in range(2, wss.max_row + 1):
                    cell = wss.cell(r, 4)
                    if str(cell.value or "") == old_name:
                        cell.value = keep_name
                        refs_updated += 1

    wb.save(EXCEL_FILE)
    clear_cache()
    log_action("MERGE_CONCAT", sheet, keep_id,
                 f"merged {drop_id} into {keep_id}; fields concatenated: "
                 f"{','.join(fields_merged) or '(none)'}; refs updated: {refs_updated}")
    return {
        "ok": True,
        "keep_id": keep_id,
        "drop_id": drop_id,
        "fields_merged": fields_merged,
        "refs_updated": refs_updated,
    }
