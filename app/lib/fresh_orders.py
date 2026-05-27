"""
Modulo per la gestione degli Ordini Fresco.

Tabelle DB:
  fresh_weeks         - una riga per ogni settimana caricata
  fresh_availability  - prodotti disponibili dal fornitore (da file Excel)
  fresh_orders_fresco - ordini assegnati ai clienti
"""
from __future__ import annotations
from datetime import datetime
from typing import Optional
import io
import pandas as pd

from lib.db import get_conn, DATABASE_URL


def _sql(query: str) -> str:
    """Converte placeholder ? → %s per PostgreSQL, li lascia ? per SQLite."""
    if DATABASE_URL:
        return query.replace("?", "%s")
    return query

# Giorni della settimana
DAYS = ["lun", "mar", "mer", "gio", "ven", "sab", "dom"]
DAYS_LABEL = ["Lunedì", "Martedì", "Mercoledì", "Giovedì", "Venerdì", "Sabato", "Domenica"]
DAY_COL = {d: f"{d}_qty" for d in DAYS}

# Categorie prodotto (righe totale nel file PREVISION)
CATEGORY_TOTALS = {"TOTAL JAMONES", "TOTAL PALETAS", "TOTAL CINTAS",
                   "TOTAL AGUJAS", "TOTAL COSTILLAS", "TOTAL SOLOMILLOS", "TOTAL RECORTES"}


def _pk_auto() -> str:
    return "SERIAL PRIMARY KEY" if DATABASE_URL else "INTEGER PRIMARY KEY AUTOINCREMENT"


def init_fresh_tables() -> None:
    """Crea le tabelle se non esistono. Idempotente."""
    pk = _pk_auto()
    num = "NUMERIC" if DATABASE_URL else "REAL"
    with get_conn() as conn:
        conn.execute(f"""
            CREATE TABLE IF NOT EXISTS fresh_weeks (
                id {pk},
                week_number INTEGER,
                year INTEGER,
                supplier TEXT,
                upload_date TEXT,
                notes TEXT
            )
        """)
        conn.execute(f"""
            CREATE TABLE IF NOT EXISTS fresh_availability (
                id {pk},
                week_id INTEGER,
                product_code TEXT,
                product_name TEXT,
                category TEXT,
                format TEXT,
                plant TEXT,
                um TEXT,
                lun_qty {num} DEFAULT 0,
                mar_qty {num} DEFAULT 0,
                mer_qty {num} DEFAULT 0,
                gio_qty {num} DEFAULT 0,
                ven_qty {num} DEFAULT 0,
                sab_qty {num} DEFAULT 0,
                dom_qty {num} DEFAULT 0
            )
        """)
        conn.execute(f"""
            CREATE TABLE IF NOT EXISTS fresh_orders_fresco (
                id {pk},
                week_id INTEGER,
                client TEXT,
                product_code TEXT,
                product_name TEXT,
                product_type TEXT,
                load_day TEXT,
                load_date TEXT,
                quantity {num} DEFAULT 0,
                um TEXT DEFAULT 'UN',
                price {num},
                delivery_notes TEXT,
                status TEXT DEFAULT 'bozza',
                created_at TEXT,
                created_by TEXT
            )
        """)


# -----------------------------------------------------------------------
# Settimane
# -----------------------------------------------------------------------
def list_weeks() -> pd.DataFrame:
    init_fresh_tables()
    with get_conn() as conn:
        raw = conn._conn if hasattr(conn, "_conn") else conn
        return pd.read_sql_query(
            "SELECT * FROM fresh_weeks ORDER BY id DESC", raw
        )


def create_week(week_number: int, year: int, supplier: str = "", notes: str = "") -> int:
    """Crea una nuova settimana. Ritorna l'id."""
    init_fresh_tables()
    ts = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    with get_conn() as conn:
        if DATABASE_URL:
            cur = conn.execute(
                _sql("INSERT INTO fresh_weeks (week_number, year, supplier, upload_date, notes) "
                "VALUES (?, ?, ?, ?, ?) RETURNING id"),
                (week_number, year, supplier, ts, notes)
            )
            return cur.fetchone()[0]
        else:
            conn.execute(
                _sql("INSERT INTO fresh_weeks (week_number, year, supplier, upload_date, notes) "
                "VALUES (?, ?, ?, ?, ?)"),
                (week_number, year, supplier, ts, notes)
            )
            raw = conn
            cur2 = raw.execute("SELECT last_insert_rowid()")
            return cur2.fetchone()[0]


def delete_week(week_id: int) -> None:
    with get_conn() as conn:
        conn.execute(_sql("DELETE FROM fresh_availability WHERE week_id = ?"), (week_id,))
        conn.execute(_sql("DELETE FROM fresh_orders_fresco WHERE week_id = ?"), (week_id,))
        conn.execute(_sql("DELETE FROM fresh_weeks WHERE id = ?"), (week_id,))


# -----------------------------------------------------------------------
# Disponibilità (da file Excel fornitore)
# -----------------------------------------------------------------------
def parse_prevision_excel(file_bytes: bytes) -> pd.DataFrame:
    """
    Legge il file PREVISION del fornitore.
    Restituisce DataFrame con colonne:
      product_code, product_name, category, format, plant, um,
      lun_qty, mar_qty, mer_qty, gio_qty, ven_qty, sab_qty, dom_qty
    """
    wb_data = io.BytesIO(file_bytes)
    # Cerca il foglio con "Previsión" o "Prevision" nel nome
    import openpyxl
    wb = openpyxl.load_workbook(wb_data, data_only=True)
    sheet_name = None
    for s in wb.sheetnames:
        if "previsi" in s.lower() or "semana" in s.lower() or "week" in s.lower():
            sheet_name = s
            break
    if not sheet_name:
        sheet_name = wb.sheetnames[0]

    ws = wb[sheet_name]
    rows = list(ws.iter_rows(values_only=True))

    # Trova riga header (contiene "Código" e "Descripción")
    header_row = None
    for i, row in enumerate(rows):
        row_str = [str(v).lower() if v else "" for v in row]
        if "código" in row_str or "codigo" in row_str or "descripcion" in row_str or "descripción" in row_str:
            header_row = i
            break
    if header_row is None:
        raise ValueError("Intestazione colonne non trovata nel file Excel.")

    headers = [str(v).strip().lower() if v else "" for v in rows[header_row]]

    # Mappa colonne
    def find_col(keywords):
        for kw in keywords:
            for i, h in enumerate(headers):
                if kw in h:
                    return i
        return None

    col_code = find_col(["código", "codigo"])
    col_desc = find_col(["descripci"])
    col_fmt  = find_col(["formato"])
    col_plant = find_col(["planta"])
    col_um   = find_col(["um", " um"])
    col_mon  = find_col(["lunes"])
    col_tue  = find_col(["martes"])
    col_wed  = find_col(["miér", "mier"])
    col_thu  = find_col(["jueves"])
    col_fri  = find_col(["viernes"])
    col_sat  = find_col(["sábado", "sabado"])
    col_sun  = find_col(["domingo"])

    day_cols = [col_mon, col_tue, col_wed, col_thu, col_fri, col_sat, col_sun]

    records = []
    current_category = ""
    for row in rows[header_row + 2:]:  # salta header e riga date
        if not any(row):
            continue
        # Riga categoria totale (es. "TOTAL JAMONES")
        label = str(row[1]).strip().upper() if row[1] else ""
        if any(cat in label for cat in ["TOTAL", "JAMONES", "PALETAS", "CINTAS", "AGUJAS", "COSTILLAS", "SOLOMILLOS", "RECORTES"]):
            if "TOTAL" not in label:
                current_category = label
            continue

        code = str(row[col_code]).strip() if col_code is not None and row[col_code] else ""
        name = str(row[col_desc]).strip() if col_desc is not None and row[col_desc] else ""
        fmt  = str(row[col_fmt]).strip() if col_fmt is not None and row[col_fmt] else ""
        plant = str(row[col_plant]).strip() if col_plant is not None and row[col_plant] else ""
        um   = str(row[col_um]).strip().upper() if col_um is not None and row[col_um] else "UN"

        if not code or code in ("0", "None") or not name or name in ("0", "None"):
            continue

        def safe_qty(col):
            if col is None:
                return 0.0
            v = row[col]
            if v is None or str(v).strip() in ("", " ", "None"):
                return 0.0
            try:
                return float(v)
            except Exception:
                return 0.0

        records.append({
            "product_code": code,
            "product_name": name,
            "category": current_category,
            "format": fmt,
            "plant": plant,
            "um": um,
            "lun_qty": safe_qty(col_mon),
            "mar_qty": safe_qty(col_tue),
            "mer_qty": safe_qty(col_wed),
            "gio_qty": safe_qty(col_thu),
            "ven_qty": safe_qty(col_fri),
            "sab_qty": safe_qty(col_sat),
            "dom_qty": safe_qty(col_sun),
        })

    return pd.DataFrame(records)


def save_availability(week_id: int, df: pd.DataFrame) -> int:
    """Salva il DataFrame disponibilità. Prima cancella righe esistenti per la settimana."""
    with get_conn() as conn:
        conn.execute(_sql("DELETE FROM fresh_availability WHERE week_id = ?"), (week_id,))
        n = 0
        for _, r in df.iterrows():
            conn.execute(
                _sql("INSERT INTO fresh_availability "
                "(week_id, product_code, product_name, category, format, plant, um, "
                "lun_qty, mar_qty, mer_qty, gio_qty, ven_qty, sab_qty, dom_qty) "
                "VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)"),
                (week_id, r.get("product_code"), r.get("product_name"),
                 r.get("category"), r.get("format"), r.get("plant"), r.get("um"),
                 float(r.get("lun_qty") or 0), float(r.get("mar_qty") or 0),
                 float(r.get("mer_qty") or 0), float(r.get("gio_qty") or 0),
                 float(r.get("ven_qty") or 0), float(r.get("sab_qty") or 0),
                 float(r.get("dom_qty") or 0))
            )
            n += 1
    return n


def get_availability(week_id: int) -> pd.DataFrame:
    init_fresh_tables()
    with get_conn() as conn:
        raw = conn._conn if hasattr(conn, "_conn") else conn
        return pd.read_sql_query(
            _sql("SELECT * FROM fresh_availability WHERE week_id = ? ORDER BY category, product_name"),
            raw, params=(week_id,)
        )


# -----------------------------------------------------------------------
# Ordini clienti
# -----------------------------------------------------------------------
def add_fresh_order(week_id: int, data: dict, user: str = "system") -> int:
    init_fresh_tables()
    ts = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    with get_conn() as conn:
        conn.execute(
            _sql("INSERT INTO fresh_orders_fresco "
            "(week_id, client, product_code, product_name, product_type, "
            "load_day, load_date, quantity, um, price, delivery_notes, status, created_at, created_by) "
            "VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)"),
            (week_id,
             data.get("client", ""),
             data.get("product_code", ""),
             data.get("product_name", ""),
             data.get("product_type", ""),
             data.get("load_day", ""),
             data.get("load_date", ""),
             float(data.get("quantity") or 0),
             data.get("um", "UN"),
             float(data.get("price") or 0) if data.get("price") else None,
             data.get("delivery_notes", ""),
             data.get("status", "bozza"),
             ts, user)
        )


def update_fresh_order(order_id: int, data: dict) -> None:
    with get_conn() as conn:
        conn.execute(
            _sql("UPDATE fresh_orders_fresco SET "
            "client=?, product_code=?, product_name=?, product_type=?, "
            "load_day=?, load_date=?, quantity=?, um=?, price=?, delivery_notes=?, status=? "
            "WHERE id=?"),
            (data.get("client", ""),
             data.get("product_code", ""),
             data.get("product_name", ""),
             data.get("product_type", ""),
             data.get("load_day", ""),
             data.get("load_date", ""),
             float(data.get("quantity") or 0),
             data.get("um", "UN"),
             float(data.get("price") or 0) if data.get("price") else None,
             data.get("delivery_notes", ""),
             data.get("status", "bozza"),
             order_id)
        )


def delete_fresh_order(order_id: int) -> None:
    with get_conn() as conn:
        conn.execute(_sql("DELETE FROM fresh_orders_fresco WHERE id = ?"), (order_id,))


def get_fresh_orders(week_id: int) -> pd.DataFrame:
    init_fresh_tables()
    with get_conn() as conn:
        raw = conn._conn if hasattr(conn, "_conn") else conn
        return pd.read_sql_query(
            _sql("SELECT * FROM fresh_orders_fresco WHERE week_id = ? "
            "ORDER BY load_day, client"),
            raw, params=(week_id,)
        )


# -----------------------------------------------------------------------
# Riepilogo allocazioni: disponibile vs allocato per giorno
# -----------------------------------------------------------------------
def get_allocation_summary(week_id: int) -> pd.DataFrame:
    """
    Per ogni giorno: totale disponibile dal fornitore, totale allocato ai clienti, rimanente.
    """
    avail = get_availability(week_id)
    orders = get_fresh_orders(week_id)

    rows = []
    for day, label in zip(DAYS, DAYS_LABEL):
        col = f"{day}_qty"
        available = float(avail[col].sum()) if not avail.empty and col in avail.columns else 0.0

        allocated = 0.0
        if not orders.empty and "load_day" in orders.columns:
            mask = orders["load_day"].str.lower().str.startswith(day[:3])
            allocated = float(orders.loc[mask, "quantity"].sum())

        rows.append({
            "Giorno": label,
            "Disponibile": available,
            "Allocato": allocated,
            "Rimanente": available - allocated,
        })
    return pd.DataFrame(rows)


# -----------------------------------------------------------------------
# Export Excel per cliente
# -----------------------------------------------------------------------
def export_orders_excel(week_id: int, client: Optional[str] = None) -> bytes:
    """
    Genera un file Excel con gli ordini.
    Se client è specificato, filtra per quel cliente.
    Altrimenti un foglio per ogni cliente.
    """
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    orders = get_fresh_orders(week_id)
    weeks_df = list_weeks()
    week_info = weeks_df[weeks_df["id"] == week_id].iloc[0] if not weeks_df.empty else None
    week_label = f"Settimana {week_info['week_number']} - {week_info['year']}" if week_info is not None else "Ordini Fresco"

    if orders.empty:
        raise ValueError("Nessun ordine da esportare per questa settimana.")

    if client:
        orders = orders[orders["client"] == client]
        if orders.empty:
            raise ValueError(f"Nessun ordine per il cliente '{client}'.")

    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    header_fill = PatternFill("solid", fgColor="0B3D91")
    header_font = Font(color="FFFFFF", bold=True)
    thin = Side(border_style="thin", color="CCCCCC")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    clients = [client] if client else sorted(orders["client"].dropna().unique())

    for cl in clients:
        df_cl = orders[orders["client"] == cl].copy()
        ws = wb.create_sheet(title=cl[:31])  # max 31 chars per sheet name

        # Titolo
        ws.merge_cells("A1:H1")
        ws["A1"] = f"ORDINE FRESCO - {cl.upper()}"
        ws["A1"].font = Font(size=14, bold=True, color="0B3D91")
        ws["A1"].alignment = Alignment(horizontal="center")

        ws.merge_cells("A2:H2")
        ws["A2"] = week_label
        ws["A2"].font = Font(size=11, color="444444")
        ws["A2"].alignment = Alignment(horizontal="center")

        # Header colonne
        cols = ["Prodotto", "Tipo/Formato", "Giorno Carico", "Data Carico",
                "Quantità", "UM", "Prezzo", "Note Scarico"]
        for j, col in enumerate(cols, 1):
            cell = ws.cell(row=4, column=j, value=col)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center")
            cell.border = border

        # Dati
        for i, (_, row) in enumerate(df_cl.iterrows(), 5):
            vals = [
                row.get("product_name", ""),
                row.get("product_type", ""),
                row.get("load_day", "").capitalize(),
                row.get("load_date", ""),
                row.get("quantity") or 0,
                row.get("um", "UN"),
                row.get("price") or "",
                row.get("delivery_notes", ""),
            ]
            for j, v in enumerate(vals, 1):
                cell = ws.cell(row=i, column=j, value=v)
                cell.border = border
                if j == 5:  # Quantità
                    cell.alignment = Alignment(horizontal="right")
                if j == 7 and v:  # Prezzo
                    cell.number_format = "#,##0.00"

        # Larghezze colonne
        widths = [35, 25, 15, 15, 12, 8, 12, 45]
        for j, w in enumerate(widths, 1):
            ws.column_dimensions[get_column_letter(j)].width = w

    out = io.BytesIO()
    wb.save(out)
    return out.getvalue()
